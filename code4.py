"""
╔══════════════════════════════════════════════════════════════╗
║               Expiro  —  Retail Expiry Manager               ║
║            Streamlit · Pandas · Plotly · openpyxl            ║
╠══════════════════════════════════════════════════════════════╣
║  Install : pip install streamlit pandas plotly openpyxl      ║
║  Run     : streamlit run app.py                              ║
╚══════════════════════════════════════════════════════════════╝
"""

import io
import os
from datetime import date, timedelta

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────────
# CONSTANTS
# ───────────────────────────────────────────────────────────────
CSV_PATH           = "inventory.csv"
CRITICAL_DAYS      = 30
EXPIRING_SOON_DAYS = 60

CATEGORIES = [
    "Dairy","Beverages","Bakery","Dry Goods",
    "Canned","Snacks","Cereals","Deli","Frozen","Other",
]
STATUS_ORDER = ["Expired","Critical","Expiring Soon","Safe"]

# ── Dark palette ──
CD = {
    "bg":        "#080d16",
    "surface":   "#0e1520",
    "surface2":  "#131c2e",
    "surface3":  "#192438",
    "border":    "#1e2d42",
    "border2":   "#243550",
    "accent":    "#f5a623",
    "accent2":   "#f97316",
    "safe":      "#10b981",
    "soon":      "#f59e0b",
    "critical":  "#ef4444",
    "expired":   "#d946ef",
    "text":      "#e2e8f0",
    "text2":     "#94a3b8",
    "muted":     "#475569",
    "safe_bg":   "#022c22",
    "soon_bg":   "#3d1f00",
    "crit_bg":   "#3d0b0b",
    "exp_bg":    "#2d0536",
    "sidebar_bg":"#0e1520",
    "input_bg":  "#131c2e",
}

# ── Light palette (soft pastels, warm off-white base) ──
CL = {
    "bg":        "#f5f1ec",
    "surface":   "#ede8e1",
    "surface2":  "#e6dfd7",
    "surface3":  "#ddd5cb",
    "border":    "#cdc6bc",
    "border2":   "#bcb4a8",
    "accent":    "#c4682a",
    "accent2":   "#a8521e",
    "safe":      "#1a7a52",
    "soon":      "#a06a10",
    "critical":  "#b83232",
    "expired":   "#8a2faa",
    "text":      "#2a2520",
    "text2":     "#4a443c",
    "muted":     "#7a7268",
    "safe_bg":   "#cceedd",
    "soon_bg":   "#f5e4cc",
    "crit_bg":   "#f5d4d4",
    "exp_bg":    "#e8d4f5",
    "sidebar_bg":"#ede8e1",
    "input_bg":  "#e6dfd7",
}

STATUS_ICON = {"Safe":"✦","Expiring Soon":"◈","Critical":"◉","Expired":"✖"}

PLOTLY_COLORS_DARK  = [CD["accent"],CD["safe"],CD["soon"],CD["critical"],CD["expired"],
                        "#a78bfa","#38bdf8","#fb7185","#4ade80","#fbbf24"]
PLOTLY_COLORS_LIGHT = [CL["accent"],CL["safe"],CL["soon"],CL["critical"],CL["expired"],
                        "#7c5cbf","#2490b5","#c0324a","#2d8a4a","#c08020"]

# ───────────────────────────────────────────────────────────────
# SESSION STATE
# ───────────────────────────────────────────────────────────────
if "data"       not in st.session_state: st.session_state.data       = None
if "edit_idx"   not in st.session_state: st.session_state.edit_idx   = None
if "light_mode" not in st.session_state: st.session_state.light_mode = False

# ── Resolve active palette ──
C  = CL if st.session_state.light_mode else CD
SCLR = {
    "Safe":          C["safe"],
    "Expiring Soon": C["soon"],
    "Critical":      C["critical"],
    "Expired":       C["expired"],
}
SBG = {
    "Safe":          C["safe_bg"],
    "Expiring Soon": C["soon_bg"],
    "Critical":      C["crit_bg"],
    "Expired":       C["exp_bg"],
}
PLOTLY_COLORS = PLOTLY_COLORS_LIGHT if st.session_state.light_mode else PLOTLY_COLORS_DARK

# ───────────────────────────────────────────────────────────────
# PLOTLY THEME
# ───────────────────────────────────────────────────────────────
def _fig_base(height=340):
    return dict(
        height=height,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", color=C["text2"], size=12),
        margin=dict(t=18, b=18, l=18, r=18),
        legend=dict(
            bgcolor="rgba(0,0,0,0)",
            bordercolor=C["border2"],
            borderwidth=1,
            font=dict(color=C["text2"], size=11),
        ),
        xaxis=dict(
            gridcolor=C["border"],
            tickcolor=C["muted"],
            linecolor=C["border2"],
            tickfont=dict(color=C["text2"]),
            showgrid=True,
            zeroline=False,
        ),
        yaxis=dict(
            gridcolor=C["border"],
            tickcolor=C["muted"],
            linecolor=C["border2"],
            tickfont=dict(color=C["text2"]),
            showgrid=True,
            zeroline=False,
        ),
    )

def stylefig(fig, height=340):
    fig.update_layout(**_fig_base(height))
    return fig

# ───────────────────────────────────────────────────────────────
# PAGE CONFIG
# ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Expiro",
    page_icon="🛡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ───────────────────────────────────────────────────────────────
# CSS
# ───────────────────────────────────────────────────────────────
is_light = st.session_state.light_mode

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap');

*, *::before, *::after {{ box-sizing: border-box; }}

html, body,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewContainer"] > .main,
section[data-testid="stSidebar"] + div {{
    background: {C["bg"]} !important;
    color: {C["text"]};
    font-family: 'Inter', sans-serif;
}}

/* ── Hide Streamlit chrome ── */
#MainMenu, footer, header, [data-testid="stDecoration"] {{ display:none !important; }}
[data-testid="stToolbar"] {{ display: none; }}

/* ── Main content padding ── */
.block-container {{ padding: 1.8rem 2.2rem 3rem !important; max-width: 1440px; }}

/* ── Sidebar ── */
[data-testid="stSidebar"] {{
    background: {C["sidebar_bg"]} !important;
    border-right: 1px solid {C["border2"]};
}}
[data-testid="stSidebar"] * {{ color: {C["text"]} !important; }}
[data-testid="stSidebar"] .stRadio > div {{ gap: 2px !important; }}
[data-testid="stSidebar"] .stRadio label {{
    padding: 10px 16px !important; border-radius: 10px !important;
    font-size: 0.88rem !important; font-weight: 500 !important;
    color: {C["text2"]} !important; cursor: pointer;
    transition: all 0.18s ease !important;
    border: 1px solid transparent !important;
}}
[data-testid="stSidebar"] .stRadio label:hover {{
    background: {C["surface2"]} !important;
    color: {C["accent"]} !important;
    border-color: {C["border2"]} !important;
}}

/* ── Floating sidebar-restore button ── */
.sidebar-restore-btn {{
    position: fixed;
    left: 0;
    top: 50%;
    transform: translateY(-50%);
    z-index: 999999;
    background: {C["accent"]};
    color: {C["bg"]};
    border: none;
    border-radius: 0 10px 10px 0;
    padding: 14px 10px;
    cursor: pointer;
    font-size: 1.1rem;
    font-weight: 900;
    box-shadow: 3px 0 16px {C["accent"]}44;
    transition: all 0.2s ease;
    writing-mode: horizontal-tb;
    line-height: 1;
}}
.sidebar-restore-btn:hover {{
    padding-left: 14px;
    box-shadow: 4px 0 20px {C["accent"]}66;
}}

/* ── Brand ── */
.eg-brand {{
    padding: 6px 4px 4px;
    border-bottom: 1px solid {C["border2"]};
    margin-bottom: 18px;
}}
.eg-name {{
    font-size: 1.55rem; font-weight: 900; letter-spacing:-1px;
    background: linear-gradient(90deg, {C["accent"]}, {C["accent2"]});
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text;
}}
.eg-tag {{
    font-size: 0.68rem; color: {C["muted"]}; text-transform: uppercase;
    letter-spacing: 0.14em; margin-top: 1px;
}}
.eg-stat {{
    font-size: 0.78rem; color: {C["text2"]};
    padding: 10px 14px;
    background: {C["surface2"]};
    border: 1px solid {C["border"]};
    border-radius: 10px;
    margin-top: 12px;
    line-height: 1.7;
}}
.eg-stat b {{ color: {C["accent"]}; }}

/* ── Page header ── */
.ph {{
    background: linear-gradient(130deg, {C["surface2"]} 0%, {C["surface3"]} 55%, {C["bg"]} 100%);
    border: 1px solid {C["border2"]};
    border-radius: 20px;
    padding: 32px 36px 26px;
    margin-bottom: 28px;
    position: relative; overflow: hidden;
}}
.ph::before {{
    content:''; position:absolute; top:0; left:0; right:0; height:3px;
    background: linear-gradient(90deg,{C["accent"]},{C["accent2"]},{C["critical"]},{C["expired"]},#a78bfa);
    border-radius: 20px 20px 0 0;
}}
.ph-title {{
    font-size: 2rem; font-weight: 900; letter-spacing:-0.8px;
    color: {C["text"]} !important; margin:0 0 5px;
}}
.ph-sub {{ font-size:0.85rem; color:{C["muted"]}; margin:0; }}

/* ── KPI cards ── */
.kpi {{
    background: {C["surface"]};
    border: 1px solid {C["border2"]};
    border-radius: 16px;
    padding: 22px 20px 20px;
    position: relative; overflow: hidden;
    min-height: 118px;
    transition: border-color 0.2s, transform 0.18s, box-shadow 0.18s;
}}
.kpi:hover {{
    transform: translateY(-3px);
    box-shadow: 0 8px 28px rgba(0,0,0,0.15);
}}
.kpi-glow {{
    position: absolute; top:-30px; right:-20px;
    width:90px; height:90px; border-radius:50%;
    opacity: 0.1; pointer-events:none;
}}
.kpi-lbl {{
    font-size:0.68rem; font-weight:700; text-transform:uppercase;
    letter-spacing:0.12em; color:{C["muted"]}; margin-bottom:10px;
}}
.kpi-val {{
    font-size:2.4rem; font-weight:900; letter-spacing:-2px; line-height:1;
}}
.kpi-sub {{ font-size:0.71rem; color:{C["muted"]}; margin-top:6px; }}
.kpi-bar {{
    position:absolute; bottom:0; left:0; right:0; height:3px;
    border-radius:0 0 16px 16px;
}}

/* ── Glass panel ── */
.glass {{
    background: {C["surface"]};
    border: 1px solid {C["border2"]};
    border-radius: 16px;
    padding: 22px 24px;
    margin-bottom: 16px;
}}
.glass-title {{
    font-size:0.72rem; font-weight:700; text-transform:uppercase;
    letter-spacing:0.12em; color:{C["muted"]}; margin-bottom:16px;
    display:flex; align-items:center; gap:8px;
}}
.glass-title::after {{
    content:''; flex:1; height:1px;
    background: linear-gradient(90deg, {C["border2"]}, transparent);
}}

/* ── Custom hbar ── */
.hbar {{ margin-bottom:13px; }}
.hbar-row {{
    display:flex; justify-content:space-between; align-items:center;
    margin-bottom:6px;
}}
.hbar-lbl {{ font-size:0.8rem; color:{C["text2"]}; font-weight:500; }}
.hbar-val {{ font-size:0.78rem; color:{C["text2"]}; font-family:'JetBrains Mono',monospace; font-weight:600; }}
.hbar-track {{
    background:{C["surface3"]}; border-radius:99px;
    height:8px; overflow:hidden;
}}
.hbar-fill {{ height:100%; border-radius:99px; transition:width .7s cubic-bezier(.4,0,.2,1); }}

/* ── Alert cards ── */
.ac {{
    border-radius:12px; padding:14px 18px;
    margin:5px 0; border:1px solid;
    display:grid; grid-template-columns:auto 1fr auto;
    align-items:center; gap:12px;
    transition: transform 0.15s;
}}
.ac:hover {{ transform: translateX(3px); }}
.ac-icon {{ font-size:1.3rem; }}
.ac-title {{ font-weight:700; font-size:0.9rem; margin-bottom:2px; color:{C["text"]}; }}
.ac-body {{
    font-size:0.76rem; color:{C["text2"]}; opacity:0.85;
    font-family:'JetBrains Mono',monospace;
}}
.ac-badge {{
    font-size:0.7rem; font-weight:700; padding:3px 10px;
    border-radius:99px; border:1px solid;
    white-space:nowrap; align-self:flex-start;
}}
.ac-expired  {{ background:{C["exp_bg"]}; border-color:{C["expired"]}44; color:{C["expired"]}; }}
.ac-critical {{ background:{C["crit_bg"]}; border-color:{C["critical"]}44; color:{C["critical"]}; }}
.ac-soon     {{ background:{C["soon_bg"]}; border-color:{C["soon"]}55; color:{C["soon"]}; }}
.ac-ok       {{ background:{C["safe_bg"]}; border-color:{C["safe"]}44; color:{C["safe"]}; }}

/* ── Status pill ── */
.pill {{
    display:inline-block; padding:2px 10px; border-radius:99px;
    font-size:0.7rem; font-weight:700; letter-spacing:0.05em;
    border:1px solid;
}}
.pill-safe     {{ background:{C["safe_bg"]};  color:{C["safe"]};     border-color:{C["safe"]}55;     }}
.pill-soon     {{ background:{C["soon_bg"]};  color:{C["soon"]};     border-color:{C["soon"]}55;     }}
.pill-critical {{ background:{C["crit_bg"]}; color:{C["critical"]}; border-color:{C["critical"]}55; }}
.pill-expired  {{ background:{C["exp_bg"]};  color:{C["expired"]};  border-color:{C["expired"]}55;  }}

/* ── Product card ── */
.pcard {{
    background:{C["surface"]}; border:1px solid {C["border2"]};
    border-radius:12px; padding:14px 18px; margin:5px 0;
    transition: border-color 0.2s, transform 0.15s, box-shadow 0.15s;
}}
.pcard:hover {{
    border-color:{C["accent"]}55;
    transform:translateY(-1px);
    box-shadow: 0 4px 16px rgba(0,0,0,0.1);
}}
.pcard-name {{ font-weight:700; font-size:0.92rem; color:{C["text"]}; }}
.pcard-meta {{ font-size:0.76rem; color:{C["muted"]}; margin-top:3px;
               font-family:'JetBrains Mono',monospace; }}

/* ── Risk banner ── */
.risk-banner {{
    background: linear-gradient(135deg, {C["crit_bg"]} 0%, {C["exp_bg"]} 100%);
    border: 1px solid {C["critical"]}44;
    border-radius: 16px; padding: 20px 28px;
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 24px;
}}
.rb-label {{ font-size:0.7rem; color:{C["critical"]}; font-weight:700;
             text-transform:uppercase; letter-spacing:.1em; margin-bottom:4px; }}
.rb-value {{ font-size:2rem; font-weight:900; color:{C["text"]}; letter-spacing:-1px; }}
.rb-pct   {{ font-size:1.5rem; font-weight:900; color:{C["critical"]}; }}
.rb-sub   {{ font-size:0.72rem; color:{C["muted"]}; margin-top:3px; }}

/* ── Streamlit widget overrides ── */
div[data-testid="metric-container"] {{
    background:{C["surface"]} !important;
    border:1px solid {C["border2"]} !important;
    border-radius:12px !important; padding:14px 18px !important;
}}
div[data-testid="metric-container"] [data-testid="metric-label"] span,
div[data-testid="metric-container"] [data-testid="metric-value"] {{
    color: {C["text"]} !important;
}}
[data-testid="stTabs"] [role="tablist"] {{
    gap: 4px; border-bottom:1px solid {C["border2"]} !important;
    background: transparent !important;
}}
[data-testid="stTabs"] [role="tab"] {{
    font-weight:600; font-size:0.83rem; color:{C["muted"]} !important;
    border:none !important; border-radius:8px 8px 0 0 !important;
    padding:8px 18px !important; transition:all 0.18s;
    background: transparent !important;
}}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {{
    color:{C["accent"]} !important;
    background:{C["surface2"]} !important;
    border-bottom:2px solid {C["accent"]} !important;
}}
[data-testid="stTabs"] [role="tab"]:hover {{ color:{C["text"]} !important; }}
[data-testid="stTabsContent"] {{
    background: transparent !important;
    border: none !important;
}}

.stButton > button {{
    background: linear-gradient(135deg,{C["accent"]},{C["accent2"]}) !important;
    color:{"#ffffff" if is_light else "#080d16"} !important; font-weight:800 !important;
    border:none !important; border-radius:10px !important;
    padding:9px 24px !important; letter-spacing:0.02em;
    transition: opacity 0.18s, transform 0.15s !important;
    box-shadow: 0 4px 18px {C["accent"]}44 !important;
}}
.stButton > button:hover {{ opacity:0.88 !important; transform:translateY(-1px) !important; }}
.stButton > button:active {{ transform:translateY(0) !important; }}

.stDownloadButton > button {{
    background:{C["surface2"]} !important;
    color:{C["accent"]} !important; font-weight:700 !important;
    border:1px solid {C["accent"]}55 !important;
    border-radius:10px !important; padding:9px 22px !important;
    transition: all 0.18s !important;
}}
.stDownloadButton > button:hover {{
    background:{C["accent"]}18 !important;
    border-color:{C["accent"]} !important;
}}

.stTextInput input, .stNumberInput input, .stDateInput input {{
    background:{C["input_bg"]} !important; color:{C["text"]} !important;
    border-color:{C["border2"]} !important; border-radius:9px !important;
    font-family:'Inter',sans-serif !important;
}}
.stSelectbox > div > div {{
    background:{C["input_bg"]} !important; color:{C["text"]} !important;
    border-color:{C["border2"]} !important; border-radius:9px !important;
}}
.stSelectbox svg {{ fill: {C["text2"]} !important; }}
.stForm {{ background:transparent !important; }}
[data-testid="stDataFrame"] {{ border-radius:12px; overflow:hidden; }}
.stDivider {{ border-color:{C["border2"]} !important; opacity:0.5; }}

/* Toggle styling */
[data-testid="stToggle"] label {{ color: {C["text2"]} !important; font-size: 0.82rem !important; }}

/* Info/success messages */
[data-testid="stAlert"] {{
    background: {C["surface2"]} !important;
    border-color: {C["border2"]} !important;
    color: {C["text"]} !important;
    border-radius: 10px !important;
}}

/* Scrollbar */
::-webkit-scrollbar {{ width:5px; height:5px; }}
::-webkit-scrollbar-track {{ background:{C["bg"]}; }}
::-webkit-scrollbar-thumb {{ background:{C["border2"]}; border-radius:3px; }}
::-webkit-scrollbar-thumb:hover {{ background:{C["accent"]}66; }}
</style>

<script>
// Sidebar restore button logic
function restoreSidebar() {{
    const collapseBtn = document.querySelector('[data-testid="collapsedControl"]');
    if (collapseBtn) {{
        collapseBtn.click();
    }}
}}
</script>

<button class="sidebar-restore-btn" onclick="restoreSidebar()" title="Open sidebar">☰</button>
""", unsafe_allow_html=True)


# ───────────────────────────────────────────────────────────────
# HTML COMPONENT HELPERS
# ───────────────────────────────────────────────────────────────

def page_header(title: str, sub: str):
    st.markdown(
        f'<div class="ph"><div class="ph-title">{title}</div>'
        f'<div class="ph-sub">{sub}</div></div>',
        unsafe_allow_html=True,
    )


def glass_title(label: str):
    st.markdown(f'<div class="glass-title">{label}</div>', unsafe_allow_html=True)


def kpi(label, value, sub, color, bar_color=None):
    bc = bar_color or color
    st.markdown(
        f'<div class="kpi">'
        f'  <div class="kpi-glow" style="background:radial-gradient({color},transparent)"></div>'
        f'  <div class="kpi-lbl">{label}</div>'
        f'  <div class="kpi-val" style="color:{color}">{value}</div>'
        f'  <div class="kpi-sub">{sub}</div>'
        f'  <div class="kpi-bar" style="background:linear-gradient(90deg,{bc},{bc}88)"></div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def hbar(label, value, max_val, color, fmt="{:.0f}", suffix=""):
    pct = min(100, round(value / max(max_val, 1) * 100))
    disp = fmt.format(value) + suffix
    st.markdown(
        f'<div class="hbar">'
        f'  <div class="hbar-row">'
        f'    <span class="hbar-lbl">{label}</span>'
        f'    <span class="hbar-val">{disp}</span>'
        f'  </div>'
        f'  <div class="hbar-track">'
        f'    <div class="hbar-fill" style="width:{pct}%;background:{color}"></div>'
        f'  </div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def pill(status):
    cls = {"Safe":"safe","Expiring Soon":"soon",
           "Critical":"critical","Expired":"expired"}.get(status,"safe")
    return f'<span class="pill pill-{cls}">{STATUS_ICON[status]} {status}</span>'


def alert_card(status, name, body, badge_text):
    cls  = {"Expired":"ac-expired","Critical":"ac-critical",
            "Expiring Soon":"ac-soon","Safe":"ac-ok"}.get(status,"ac-ok")
    icon = {"Expired":"💀","Critical":"🔴","Expiring Soon":"🟡","Safe":"✅"}.get(status,"")
    c    = SCLR[status]
    st.markdown(
        f'<div class="ac {cls}">'
        f'  <div class="ac-icon">{icon}</div>'
        f'  <div>'
        f'    <div class="ac-title">{name}</div>'
        f'    <div class="ac-body">{body}</div>'
        f'  </div>'
        f'  <div class="ac-badge" style="color:{c};border-color:{c}55;background:{SBG[status]}">'
        f'    {badge_text}'
        f'  </div>'
        f'</div>',
        unsafe_allow_html=True,
    )


# ───────────────────────────────────────────────────────────────
# DATA LAYER
# ───────────────────────────────────────────────────────────────

def sample_data() -> pd.DataFrame:
    t = date.today()
    rows = [
        ("Whole Milk 1L",       "Dairy",     "B001", 50,  t-timedelta(10),  t+timedelta(5),   45),
        ("Cheddar Cheese 500g", "Dairy",     "B002", 30,  t-timedelta(20),  t+timedelta(22),  130),
        ("Orange Juice 2L",     "Beverages", "B003", 40,  t-timedelta(30),  t+timedelta(45),  95),
        ("White Bread",         "Bakery",    "B004", 25,  t-timedelta(3),   t+timedelta(3),   40),
        ("Pasta 500g",          "Dry Goods", "B005", 100, t-timedelta(60),  t+timedelta(180), 30),
        ("Tomato Sauce 400g",   "Canned",    "B006", 80,  t-timedelta(90),  t+timedelta(120), 35),
        ("Greek Yogurt 200g",   "Dairy",     "B007", 20,  t-timedelta(5),   t-timedelta(1),   28),
        ("Sparkling Water 1L",  "Beverages", "B008", 60,  t-timedelta(120), t+timedelta(90),  18),
        ("Granola Bars x6",     "Snacks",    "B009", 45,  t-timedelta(45),  t+timedelta(50),  75),
        ("Butter 250g",         "Dairy",     "B010", 18,  t-timedelta(15),  t+timedelta(18),  62),
        ("Corn Flakes 500g",    "Cereals",   "B011", 35,  t-timedelta(80),  t+timedelta(200), 80),
        ("Salted Crackers",     "Snacks",    "B012", 55,  t-timedelta(60),  t+timedelta(40),  42),
        ("Apple Juice 1L",      "Beverages", "B013", 28,  t-timedelta(40),  t-timedelta(5),   55),
        ("Sliced Ham 200g",     "Deli",      "B014", 12,  t-timedelta(8),   t+timedelta(10),  110),
        ("Cream Cheese 150g",   "Dairy",     "B015", 22,  t-timedelta(12),  t+timedelta(14),  85),
        ("Frozen Pizza",        "Frozen",    "B016", 15,  t-timedelta(30),  t+timedelta(25),  220),
        ("Biscuits 200g",       "Snacks",    "B017", 70,  t-timedelta(90),  t+timedelta(150), 55),
        ("Skim Milk 500ml",     "Dairy",     "B018", 40,  t-timedelta(7),   t+timedelta(8),   38),
    ]
    df = pd.DataFrame(rows, columns=[
        "Product Name","Category","Batch Number","Quantity",
        "Manufacturing Date","Expiry Date","Cost Per Unit",
    ])
    df["Manufacturing Date"] = pd.to_datetime(df["Manufacturing Date"]).dt.date
    df["Expiry Date"]        = pd.to_datetime(df["Expiry Date"]).dt.date
    return df


def load_data() -> pd.DataFrame:
    if os.path.exists(CSV_PATH):
        df = pd.read_csv(CSV_PATH)
        df["Manufacturing Date"] = pd.to_datetime(df["Manufacturing Date"]).dt.date
        df["Expiry Date"]        = pd.to_datetime(df["Expiry Date"]).dt.date
        df["Cost Per Unit"]      = pd.to_numeric(df["Cost Per Unit"], errors="coerce").fillna(0)
    else:
        df = sample_data()
        df.to_csv(CSV_PATH, index=False)
    return df


def save_data(df: pd.DataFrame):
    df.to_csv(CSV_PATH, index=False)


def classify(days: int) -> str:
    if days < 0:                   return "Expired"
    if days < CRITICAL_DAYS:       return "Critical"
    if days < EXPIRING_SOON_DAYS:  return "Expiring Soon"
    return "Safe"


def enrich(raw: pd.DataFrame) -> pd.DataFrame:
    d = raw.copy()
    today = date.today()
    d["Days Remaining"] = pd.to_datetime(d["Expiry Date"]).apply(
        lambda x: (x.date()-today).days
    )
    d["Status"]        = d["Days Remaining"].apply(classify)
    d["Risk Score"]    = d["Days Remaining"].apply(
        lambda x: 100 if x < 0 else max(0, round(100-(x/EXPIRING_SOON_DAYS)*100))
    )
    d["Est. Loss (₹)"] = (d["Quantity"]*d["Cost Per Unit"]).round(2)
    return d


def gdf() -> pd.DataFrame:
    return enrich(st.session_state.data)


# ───────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ───────────────────────────────────────────────────────────────

def build_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    STATUS_FILL = {
        "Safe":          PatternFill("solid", fgColor="022C22"),
        "Expiring Soon": PatternFill("solid", fgColor="3D1F00"),
        "Critical":      PatternFill("solid", fgColor="3D0B0B"),
        "Expired":       PatternFill("solid", fgColor="2D0536"),
    }
    STATUS_FONT_CLR = {
        "Safe":"10B981","Expiring Soon":"F59E0B",
        "Critical":"EF4444","Expired":"D946EF",
    }
    thin   = Side(style="thin", color="1E2D42")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        "Product Name","Category","Batch Number","Quantity",
        "Manufacturing Date","Expiry Date","Cost Per Unit",
        "Days Remaining","Status","Risk Score","Est. Loss (₹)",
    ]

    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value     = "Expiro — Inventory Report"
    tc.font      = Font(name="Calibri", bold=True, size=16, color="F5A623")
    tc.fill      = PatternFill("solid", fgColor="080D16")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:K2")
    dc = ws["A2"]
    dc.value     = f"Generated: {date.today().strftime('%d %B %Y')}"
    dc.font      = Font(name="Calibri", size=10, color="64748B")
    dc.fill      = PatternFill("solid", fgColor="0E1520")
    dc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    hf = PatternFill("solid", fgColor="131C2E")
    for ci, col in enumerate(cols, 1):
        cell           = ws.cell(row=3, column=ci, value=col)
        cell.font      = Font(name="Calibri", bold=True, size=10, color="F5A623")
        cell.fill      = hf
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 28

    display_df = df[cols].copy()
    for ri, (_, row) in enumerate(display_df.iterrows(), 4):
        status   = row["Status"]
        sfill    = STATUS_FILL.get(status, PatternFill("solid", fgColor="0E1520"))
        sfc      = STATUS_FONT_CLR.get(status, "E2E8F0")
        row_fill = (PatternFill("solid", fgColor="0E1520") if ri % 2 == 0
                    else PatternFill("solid", fgColor="131C2E"))
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=str(row[col]))
            if col == "Status":
                cell.fill = sfill
                cell.font = Font(name="Calibri", bold=True, size=10, color=sfc)
            else:
                cell.fill = row_fill
                cell.font = Font(name="Calibri", size=10, color="E2E8F0")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
        ws.row_dimensions[ri].height = 20

    widths = [28,14,14,10,18,18,16,16,16,12,16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = "Status Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="F5A623")
    ws2["A1"].fill = PatternFill("solid", fgColor="080D16")
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 30

    ws2["A2"], ws2["B2"], ws2["C2"] = "Status","Count","Est. Loss (₹)"
    for c in ["A2","B2","C2"]:
        ws2[c].font      = Font(name="Calibri", bold=True, size=10, color="F5A623")
        ws2[c].fill      = PatternFill("solid", fgColor="131C2E")
        ws2[c].alignment = Alignment(horizontal="center")
        ws2[c].border    = border

    for ri, status in enumerate(STATUS_ORDER, 3):
        sub  = df[df["Status"]==status]
        cnt  = len(sub)
        loss = float(sub["Est. Loss (₹)"].sum())
        fc   = STATUS_FONT_CLR.get(status, "E2E8F0")
        fil  = STATUS_FILL.get(status, PatternFill("solid", fgColor="0E1520"))
        for ci, val in enumerate([status, cnt, f"₹{loss:,.2f}"], 1):
            cell           = ws2.cell(ri, ci, val)
            cell.font      = Font(name="Calibri", bold=(ci==1), size=10,
                                  color=fc if ci==1 else "E2E8F0")
            cell.fill      = fil
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
        ws2.row_dimensions[ri].height = 22

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ───────────────────────────────────────────────────────────────
# LOAD DATA (after session_state check)
# ───────────────────────────────────────────────────────────────
if st.session_state.data is None:
    st.session_state.data = load_data()


# ═══════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    data_sb  = gdf()
    exp_cnt  = int((data_sb["Status"]=="Expired").sum())
    crit_cnt = int((data_sb["Status"]=="Critical").sum())

    st.markdown(
        f'<div class="eg-brand">'
        f'  <div class="eg-name">🛡 Expiro</div>'
        f'  <div class="eg-tag">Retail Expiry Manager</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    page = st.radio(
        "nav", label_visibility="collapsed",
        options=[
            "📊  Dashboard",
            "➕  Product Entry",
            "⚠️  Risk Analysis",
            "🔔  Alerts",
            "📈  Analytics",
            "🔍  Search & Filter",
        ],
    )

    # ── Theme toggle ──
    st.markdown("<br>", unsafe_allow_html=True)
    new_mode = st.toggle("☀️  Light mode", value=st.session_state.light_mode)
    if new_mode != st.session_state.light_mode:
        st.session_state.light_mode = new_mode
        st.rerun()

    # ── Status stats ──
    alert_html = ""
    if exp_cnt:
        alert_html += f'<span style="color:{C["expired"]};font-weight:700">✖ {exp_cnt} expired</span>  '
    if crit_cnt:
        alert_html += f'<span style="color:{C["critical"]};font-weight:700">◉ {crit_cnt} critical</span>'

    st.markdown(
        f'<div class="eg-stat">'
        f'  📅 <b>{date.today().strftime("%d %b %Y")}</b><br>'
        f'  📦 <b>{len(st.session_state.data)}</b> products tracked'
        f'  {"<br>" + alert_html if alert_html else ""}'
        f'</div>',
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("⟳  Refresh Data", use_container_width=True):
        st.session_state.data = load_data()
        st.success("Refreshed!")


# ═══════════════════════════════════════════════════════════════
# PAGE 1 ── DASHBOARD
# ═══════════════════════════════════════════════════════════════
if page == "📊  Dashboard":
    page_header(
        "Inventory Dashboard",
        f"Live expiry overview · {date.today().strftime('%A, %d %B %Y')}",
    )
    data     = gdf()
    total    = len(data)
    safe     = int((data["Status"]=="Safe").sum())
    soon     = int((data["Status"]=="Expiring Soon").sum())
    critical = int((data["Status"]=="Critical").sum())
    expired  = int((data["Status"]=="Expired").sum())
    total_val = float(data["Est. Loss (₹)"].sum())
    risk_val  = float(data[data["Status"].isin(["Expired","Critical","Expiring Soon"])]["Est. Loss (₹)"].sum())
    risk_pct  = round(risk_val / max(total_val, 1) * 100)

    k0,k1,k2,k3,k4 = st.columns(5)
    with k0: kpi("Total Products",  total,    "all SKUs tracked",  C["accent"],   C["accent2"])
    with k1: kpi("Safe",            safe,     "> 60 days left",    C["safe"])
    with k2: kpi("Expiring Soon",   soon,     "30 – 60 days",      C["soon"])
    with k3: kpi("Critical",        critical, "< 30 days",         C["critical"])
    with k4: kpi("Expired",         expired,  "immediate action",  C["expired"])

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown(
        f'<div class="risk-banner">'
        f'  <div>'
        f'    <div class="rb-label">Estimated Inventory Value at Risk</div>'
        f'    <div class="rb-value">₹{risk_val:,.2f}</div>'
        f'    <div class="rb-sub">{expired} expired · {critical} critical · {soon} expiring soon</div>'
        f'  </div>'
        f'  <div style="text-align:right">'
        f'    <div class="rb-sub">% of total inventory value</div>'
        f'    <div class="rb-pct">{risk_pct}%</div>'
        f'  </div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    r2a, r2b = st.columns([1, 1.4])

    with r2a:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Stock Health Overview")
        status_counts = {s: int((data["Status"]==s).sum()) for s in STATUS_ORDER}
        fig_donut = go.Figure(go.Pie(
            labels=list(status_counts.keys()),
            values=list(status_counts.values()),
            hole=0.62,
            marker=dict(
                colors=[SCLR[s] for s in STATUS_ORDER],
                line=dict(color=C["bg"], width=3),
            ),
            textinfo="percent",
            textfont=dict(size=12, color=C["text"]),
            hovertemplate="<b>%{label}</b><br>%{value} products<br>%{percent}<extra></extra>",
        ))
        fig_donut.add_annotation(
            text=f"<b>{total}</b><br><span style='font-size:10px'>products</span>",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=18, color=C["text"]),
            align="center",
        )
        fig_donut.update_layout(**_fig_base(300))
        fig_donut.update_layout(
            showlegend=True,
            legend=dict(orientation="v", x=1.02, y=0.5,
                        font=dict(color=C["text2"],size=11)),
            margin=dict(t=10,b=10,l=10,r=10),
        )
        st.plotly_chart(fig_donut, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    with r2b:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Products per Category")
        cat_status = (
            data.groupby(["Category","Status"])
            .size().reset_index(name="Count")
        )
        fig_bar = px.bar(
            cat_status, x="Category", y="Count", color="Status",
            color_discrete_map=SCLR,
            category_orders={"Status": STATUS_ORDER},
            barmode="stack",
        )
        fig_bar.update_traces(
            marker_line_width=0,
            hovertemplate="<b>%{x}</b><br>%{customdata[0]}: %{y}<extra></extra>",
            customdata=cat_status[["Status"]].values,
        )
        fig_bar.update_layout(**_fig_base(300))
        fig_bar.update_layout(
            xaxis_tickangle=-30,
            legend_title_text="",
            bargap=0.35,
            margin=dict(t=10,b=10,l=10,r=10),
        )
        st.plotly_chart(fig_bar, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    r3a, r3b = st.columns([1.5, 1])

    with r3a:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Top 10 Closest to Expiry")
        top10 = (
            data[data["Days Remaining"] >= 0]
            .sort_values("Days Remaining").head(10).copy()
        )
        if not top10.empty:
            top10["Label"] = top10["Product Name"].str[:24]
            fig_h = px.bar(
                top10.sort_values("Days Remaining", ascending=True),
                x="Days Remaining", y="Label",
                orientation="h",
                color="Status",
                color_discrete_map=SCLR,
                text="Days Remaining",
                category_orders={"Status": STATUS_ORDER},
                custom_data=["Product Name","Category","Quantity"],
            )
            fig_h.update_traces(
                texttemplate="%{text}d",
                textposition="outside",
                textfont=dict(color=C["text2"], size=10),
                marker_line_width=0,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Category: %{customdata[1]}<br>"
                    "Qty: %{customdata[2]}<br>"
                    "Days left: %{x}<extra></extra>"
                ),
            )
            fig_h.update_layout(**_fig_base(320))
            fig_h.update_layout(
                xaxis_title="Days Remaining",
                yaxis_title="",
                showlegend=True,
                legend_title_text="",
                margin=dict(t=10,b=10,l=10,r=60),
            )
            st.plotly_chart(fig_h, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    with r3b:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Financial Risk by Status")
        risk_df = (data.groupby("Status")["Est. Loss (₹)"]
                   .sum().reindex(STATUS_ORDER).fillna(0).reset_index())
        fig_risk = px.bar(
            risk_df, x="Status", y="Est. Loss (₹)",
            color="Status", color_discrete_map=SCLR,
            text="Est. Loss (₹)",
            category_orders={"Status": STATUS_ORDER},
        )
        fig_risk.update_traces(
            texttemplate="₹%{text:,.0f}",
            textposition="outside",
            textfont=dict(color=C["text2"],size=9),
            marker_line_width=0,
        )
        fig_risk.update_layout(**_fig_base(320))
        fig_risk.update_layout(
            showlegend=False, bargap=0.4,
            margin=dict(t=10,b=10,l=10,r=10),
            yaxis_title="Estimated Loss (₹)",
        )
        st.plotly_chart(fig_risk, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# PAGE 2 ── PRODUCT ENTRY
# ═══════════════════════════════════════════════════════════════
elif page == "➕  Product Entry":
    page_header("Product Entry", "Add new stock or manage existing inventory records")

    tab_add, tab_manage = st.tabs(["  ✦  Add New Product  ","  ✦  Manage Products  "])

    with tab_add:
        st.markdown("<br>", unsafe_allow_html=True)
        with st.form("add_form", clear_on_submit=True):
            st.markdown('<div class="glass-title">Product Details</div>', unsafe_allow_html=True)
            c1,c2 = st.columns(2)
            name  = c1.text_input("Product Name *", placeholder="e.g. Full Cream Milk 500ml")
            cat   = c2.selectbox("Category *", CATEGORIES)
            c3,c4 = st.columns(2)
            batch = c3.text_input("Batch Number *", placeholder="e.g. B0042")
            qty   = c4.number_input("Quantity *", min_value=1, step=1, value=10)
            c5,c6,c7 = st.columns(3)
            mfg  = c5.date_input("Manufacturing Date *", value=date.today()-timedelta(30))
            exp  = c6.date_input("Expiry Date *",        value=date.today()+timedelta(90))
            cost = c7.number_input("Cost Per Unit (₹) *", min_value=0.0, step=5.0,
                                   value=50.0, format="%.2f")
            st.markdown("<br>", unsafe_allow_html=True)
            sub = st.form_submit_button("＋  Add to Inventory", use_container_width=True)

        if sub:
            errs = []
            if not name.strip():  errs.append("Product Name is required.")
            if not batch.strip(): errs.append("Batch Number is required.")
            if exp <= mfg:        errs.append("Expiry date must be after Manufacturing date.")
            if errs:
                for e in errs: st.error(e)
            else:
                new = pd.DataFrame([{
                    "Product Name": name.strip(), "Category": cat,
                    "Batch Number": batch.strip(), "Quantity": int(qty),
                    "Manufacturing Date": mfg, "Expiry Date": exp,
                    "Cost Per Unit": float(cost),
                }])
                st.session_state.data = pd.concat(
                    [st.session_state.data, new], ignore_index=True)
                save_data(st.session_state.data)
                st.success(f"✅  '{name}' added successfully!")

    with tab_manage:
        raw = gdf()
        if raw.empty:
            st.info("No products found. Add one in the 'Add New Product' tab.")
        else:
            srch = st.text_input("Filter products", placeholder="Search by name…",
                                 label_visibility="collapsed")
            disp = (raw[raw["Product Name"].str.contains(srch, case=False, na=False)]
                    if srch else raw)

            for i, row in disp.iterrows():
                cc, ce, cd = st.columns([8,1,1])
                with cc:
                    p_html = pill(row["Status"])
                    clr    = SCLR.get(row["Status"], C["muted"])
                    st.markdown(
                        f'<div class="pcard">'
                        f'  <div class="pcard-name">{row["Product Name"]} &nbsp;{p_html}</div>'
                        f'  <div class="pcard-meta">'
                        f'    {row["Category"]} · Batch: {row["Batch Number"]} · '
                        f'    Qty: {row["Quantity"]} · Expires: {row["Expiry Date"]} · '
                        f'    <span style="color:{clr};font-weight:600">{row["Days Remaining"]}d left</span>'
                        f'  </div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                with ce:
                    st.markdown("<div style='margin-top:14px'></div>", unsafe_allow_html=True)
                    if st.button("✏", key=f"ed_{i}"):
                        st.session_state.edit_idx = i
                with cd:
                    st.markdown("<div style='margin-top:14px'></div>", unsafe_allow_html=True)
                    if st.button("✖", key=f"dl_{i}"):
                        st.session_state.data = (
                            st.session_state.data.drop(index=i).reset_index(drop=True))
                        save_data(st.session_state.data)
                        st.rerun()

            eidx = st.session_state.edit_idx
            if eidx is not None and eidx in st.session_state.data.index:
                r = st.session_state.data.loc[eidx]
                st.divider()
                st.markdown(
                    f'<div class="glass-title">Editing: {r["Product Name"]}</div>',
                    unsafe_allow_html=True,
                )
                with st.form("edit_form"):
                    ea,eb  = st.columns(2)
                    e_name = ea.text_input("Product Name", value=r["Product Name"])
                    ci     = CATEGORIES.index(r["Category"]) if r["Category"] in CATEGORIES else 0
                    e_cat  = eb.selectbox("Category", CATEGORIES, index=ci)
                    ec,ed  = st.columns(2)
                    e_b = ec.text_input("Batch Number", value=r["Batch Number"])
                    e_q = ed.number_input("Quantity", min_value=1, value=int(r["Quantity"]))
                    ee,ef,eg = st.columns(3)
                    e_m = ee.date_input("Mfg Date",    value=r["Manufacturing Date"])
                    e_e = ef.date_input("Expiry Date", value=r["Expiry Date"])
                    e_c = eg.number_input("Cost/Unit", min_value=0.0,
                                          value=float(r["Cost Per Unit"]), format="%.2f")
                    sv = st.form_submit_button("💾  Save Changes", use_container_width=True)
                if sv:
                    if e_e <= e_m:
                        st.error("Expiry date must be after Manufacturing date.")
                    else:
                        for k, v in zip(
                            ["Product Name","Category","Batch Number","Quantity",
                             "Manufacturing Date","Expiry Date","Cost Per Unit"],
                            [e_name, e_cat, e_b, e_q, e_m, e_e, e_c]
                        ):
                            st.session_state.data.at[eidx, k] = v
                        save_data(st.session_state.data)
                        st.session_state.edit_idx = None
                        st.success("Saved!")
                        st.rerun()


# ═══════════════════════════════════════════════════════════════
# PAGE 3 ── RISK ANALYSIS
# ═══════════════════════════════════════════════════════════════
elif page == "⚠️  Risk Analysis":
    page_header("Expiry Risk Analysis", "Classification, distribution, and risk breakdown at a glance")

    data = gdf()

    k1,k2,k3,k4 = st.columns(4)
    for col, status in zip([k1,k2,k3,k4], STATUS_ORDER):
        n = int((data["Status"]==status).sum())
        with col:
            kpi(status, n, f"{STATUS_ICON[status]} products", SCLR[status])

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Row 1: Days remaining histogram + risk hbars side-by-side ──
    r1a, r1b = st.columns([1.6, 1])

    with r1a:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Days Remaining Distribution")
        active = data[data["Days Remaining"].between(-30, 210)].copy()
        fig_hist = px.histogram(
            active, x="Days Remaining", color="Status",
            nbins=36, color_discrete_map=SCLR,
            category_orders={"Status": STATUS_ORDER},
            barmode="overlay",
        )
        for x_val, label, clr in [
            (0,  "Expired →",  C["expired"]),
            (30, "Critical →", C["critical"]),
            (60, "Soon →",     C["soon"]),
        ]:
            fig_hist.add_vline(
                x=x_val, line_dash="dot", line_color=clr, line_width=1.5,
                annotation_text=label,
                annotation_font=dict(color=clr, size=10),
                annotation_position="top right",
            )
        fig_hist.update_traces(opacity=0.80, marker_line_width=0)
        fig_hist.update_layout(**_fig_base(320))
        fig_hist.update_layout(
            xaxis_title="Days Remaining", yaxis_title="Products",
            legend_title_text="", bargap=0.05,
            margin=dict(t=20,b=10,l=10,r=10),
        )
        st.plotly_chart(fig_hist, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    with r1b:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Average Risk Score by Status")
        risk_avg = (
            data.groupby("Status")["Risk Score"]
            .mean().reindex(STATUS_ORDER).fillna(0).reset_index()
        )
        fig_avg = px.bar(
            risk_avg, x="Status", y="Risk Score",
            color="Status", color_discrete_map=SCLR,
            text="Risk Score",
            category_orders={"Status": STATUS_ORDER},
        )
        fig_avg.update_traces(
            texttemplate="%{text:.0f}",
            textposition="outside",
            textfont=dict(color=C["text2"], size=11),
            marker_line_width=0,
            showlegend=False,
        )
        fig_avg.update_layout(**_fig_base(320))
        fig_avg.update_layout(
            bargap=0.4, xaxis_title="", yaxis_title="Avg Risk Score (0–100)",
            yaxis_range=[0, 115],
            margin=dict(t=20,b=10,l=10,r=10),
        )
        st.plotly_chart(fig_avg, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Row 2: Risk score by category (horizontal bars) ──
    st.markdown('<div class="glass">', unsafe_allow_html=True)
    glass_title("Risk Score by Category")
    cat_risk = (
        data.groupby("Category")
        .agg(Avg_Risk=("Risk Score","mean"), At_Risk=("Status",
             lambda x: x.isin(["Critical","Expiring Soon","Expired"]).sum()),
             Products=("Product Name","count"))
        .reset_index().sort_values("Avg_Risk", ascending=True)
    )
    fig_cr = px.bar(
        cat_risk, y="Category", x="Avg_Risk",
        orientation="h",
        color="Avg_Risk",
        color_continuous_scale=[C["safe"], C["soon"], C["critical"]],
        text="Avg_Risk",
        custom_data=["Products","At_Risk"],
    )
    fig_cr.update_traces(
        texttemplate="%{text:.0f}",
        textposition="outside",
        textfont=dict(color=C["text2"], size=10),
        marker_line_width=0,
        hovertemplate=(
            "<b>%{y}</b><br>Avg Risk: %{x:.0f}<br>"
            "Products: %{customdata[0]}<br>At Risk: %{customdata[1]}<extra></extra>"
        ),
    )
    fig_cr.update_layout(**_fig_base(300))
    fig_cr.update_coloraxes(showscale=False)
    fig_cr.update_layout(
        xaxis_title="Average Risk Score", yaxis_title="",
        margin=dict(t=10, b=10, l=10, r=60),
    )
    st.plotly_chart(fig_cr, use_container_width=True, config={"displayModeBar":False})
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Per-status tables ──
    SHOW = ["Product Name","Category","Batch Number","Quantity",
            "Expiry Date","Days Remaining","Risk Score","Est. Loss (₹)"]
    st.markdown("<br>", unsafe_allow_html=True)

    for status in STATUS_ORDER:
        subset = (data[data["Status"]==status]
                  .sort_values("Days Remaining").reset_index(drop=True))
        if subset.empty: continue
        c = SCLR[status]
        st.markdown(
            f'<div class="glass-title" style="color:{c}">'
            f'{STATUS_ICON[status]} {status} &nbsp;'
            f'<span style="background:{SBG[status]};color:{c};padding:2px 10px;'
            f'border-radius:99px;font-size:0.75rem;border:1px solid {c}44">'
            f'{len(subset)} products</span></div>',
            unsafe_allow_html=True,
        )
        st.dataframe(
            subset[SHOW], use_container_width=True, hide_index=True,
            column_config={
                "Risk Score":     st.column_config.ProgressColumn(
                    "Risk Score", min_value=0, max_value=100, format="%d"),
                "Days Remaining": st.column_config.NumberColumn("Days Left", format="%d d"),
                "Est. Loss (₹)":  st.column_config.NumberColumn(format="₹%.2f"),
            },
        )
        st.markdown("<br>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# PAGE 4 ── ALERTS
# ═══════════════════════════════════════════════════════════════
elif page == "🔔  Alerts":
    page_header("Alerts Centre", "Products requiring immediate or near-term attention")

    data    = gdf()
    expired = data[data["Status"]=="Expired"].sort_values("Days Remaining")
    crit    = data[data["Status"]=="Critical"].sort_values("Days Remaining")
    soon    = data[data["Status"]=="Expiring Soon"].sort_values("Days Remaining")
    risk_v  = float(data[data["Status"].isin(
        ["Expired","Critical","Expiring Soon"])]["Est. Loss (₹)"].sum())

    t1,t2,t3,t4 = st.columns(4)
    with t1: kpi("Expired",       len(expired), "remove now",   C["expired"])
    with t2: kpi("Critical",      len(crit),    "< 30 days",    C["critical"])
    with t3: kpi("Expiring Soon", len(soon),    "30–60 days",   C["soon"])
    with t4: kpi("Total At Risk", f"₹{risk_v:,.0f}", "est. loss", C["accent"], C["accent2"])

    st.markdown("<br>", unsafe_allow_html=True)

    at_risk_all = data[data["Status"].isin(["Expired","Critical","Expiring Soon"])].copy()
    if not at_risk_all.empty:
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("At-Risk Products — Days Remaining vs Estimated Loss")
        fig_sc = px.scatter(
            at_risk_all,
            x="Days Remaining",
            y="Est. Loss (₹)",
            color="Status",
            size="Quantity",
            size_max=30,
            color_discrete_map=SCLR,
            hover_data={"Product Name":True,"Category":True,"Batch Number":True,
                        "Quantity":True,"Days Remaining":True,"Est. Loss (₹)":True},
            text="Product Name",
        )
        fig_sc.update_traces(
            textposition="top center",
            textfont=dict(size=9, color=C["text2"]),
            marker=dict(line=dict(width=1, color=C["bg"])),
        )
        fig_sc.update_layout(**_fig_base(360))
        fig_sc.update_layout(
            xaxis_title="Days Remaining",
            yaxis_title="Estimated Loss (₹)",
            legend_title_text="",
            margin=dict(t=20,b=10,l=10,r=10),
        )
        st.plotly_chart(fig_sc, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_exp, col_crit = st.columns(2)

    with col_exp:
        st.markdown(
            f'<div class="glass-title" style="color:{C["expired"]}">✖ Expired ({len(expired)})</div>',
            unsafe_allow_html=True,
        )
        if expired.empty:
            alert_card("Safe","No expired products","All clear — great job!","✦ clean")
        else:
            for _, r in expired.iterrows():
                alert_card("Expired", r["Product Name"],
                    f"Batch {r['Batch Number']}  ·  {abs(r['Days Remaining'])}d ago  ·  Qty {r['Quantity']}",
                    f"₹{r['Est. Loss (₹)']:.0f}")

    with col_crit:
        st.markdown(
            f'<div class="glass-title" style="color:{C["critical"]}">◉ Critical ({len(crit)})</div>',
            unsafe_allow_html=True,
        )
        if crit.empty:
            alert_card("Safe","No critical products","All clear!","✦ clean")
        else:
            for _, r in crit.iterrows():
                alert_card("Critical", r["Product Name"],
                    f"Batch {r['Batch Number']}  ·  {r['Days Remaining']}d left  ·  Qty {r['Quantity']}",
                    f"{r['Risk Score']}/100")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        f'<div class="glass-title" style="color:{C["soon"]}">◈ Expiring Soon ({len(soon)})</div>',
        unsafe_allow_html=True,
    )
    if soon.empty:
        alert_card("Safe","Nothing in 30–60-day window","All clear!","✦ clean")
    else:
        cols_s = st.columns(2)
        for idx2, (_, r) in enumerate(soon.iterrows()):
            with cols_s[idx2 % 2]:
                alert_card("Expiring Soon", r["Product Name"],
                    f"Batch {r['Batch Number']}  ·  {r['Days Remaining']}d  ·  Qty {r['Quantity']}",
                    f"{r['Days Remaining']}d")


# ═══════════════════════════════════════════════════════════════
# PAGE 5 ── ANALYTICS
# ═══════════════════════════════════════════════════════════════
elif page == "📈  Analytics":
    page_header("Reports & Analytics", "Deep-dive charts, financial risk, and styled Excel exports")

    data = gdf()

    tab_cat, tab_fin, tab_dl = st.tabs([
        "  📦 Category Breakdown  ",
        "  💸 Financial Risk  ",
        "  📥 Export  ",
    ])

    # ── CATEGORY BREAKDOWN ──
    with tab_cat:
        st.markdown("<br>", unsafe_allow_html=True)
        cat_df = (
            data.groupby("Category")
            .agg(
                Products=("Product Name","count"),
                Total_Qty=("Quantity","sum"),
                At_Risk=("Status", lambda x: x.isin(["Critical","Expiring Soon","Expired"]).sum()),
                Est_Loss=("Est. Loss (₹)","sum"),
            ).reset_index().rename(columns={"Est_Loss":"Est. Loss (₹)"})
            .sort_values("Products", ascending=False)
        )

        ca, cb = st.columns(2)
        with ca:
            st.markdown('<div class="glass">', unsafe_allow_html=True)
            glass_title("Safe vs At-Risk per Category")
            cat_df["Safe_Count"] = cat_df["Products"] - cat_df["At_Risk"]
            fig_cmp = go.Figure()
            fig_cmp.add_trace(go.Bar(
                name="Safe", x=cat_df["Category"], y=cat_df["Safe_Count"],
                marker=dict(color=C["safe"], line=dict(width=0)),
                hovertemplate="<b>%{x}</b><br>Safe: %{y}<extra></extra>",
            ))
            fig_cmp.add_trace(go.Bar(
                name="At Risk", x=cat_df["Category"], y=cat_df["At_Risk"],
                marker=dict(color=C["critical"], line=dict(width=0)),
                hovertemplate="<b>%{x}</b><br>At Risk: %{y}<extra></extra>",
            ))
            fig_cmp.update_layout(**_fig_base(300))
            fig_cmp.update_layout(
                barmode="stack", bargap=0.4, xaxis_tickangle=-30,
                legend=dict(orientation="h", y=-0.28, x=0.28),
                margin=dict(t=10,b=10,l=10,r=10),
            )
            st.plotly_chart(fig_cmp, use_container_width=True, config={"displayModeBar":False})
            st.markdown('</div>', unsafe_allow_html=True)

        with cb:
            st.markdown('<div class="glass">', unsafe_allow_html=True)
            glass_title("Estimated Loss per Category")
            cat_loss = cat_df.sort_values("Est. Loss (₹)", ascending=True)
            fig_cl = px.bar(
                cat_loss, y="Category", x="Est. Loss (₹)",
                orientation="h",
                color="Est. Loss (₹)",
                color_continuous_scale=[C["safe"], C["soon"], C["critical"]],
                text="Est. Loss (₹)",
            )
            fig_cl.update_traces(
                texttemplate="₹%{text:,.0f}",
                textposition="outside",
                textfont=dict(color=C["text2"], size=9),
                marker_line_width=0,
            )
            fig_cl.update_layout(**_fig_base(300))
            fig_cl.update_coloraxes(showscale=False)
            fig_cl.update_layout(
                xaxis_title="Est. Loss (₹)", yaxis_title="",
                margin=dict(t=10,b=10,l=10,r=70),
            )
            st.plotly_chart(fig_cl, use_container_width=True, config={"displayModeBar":False})
            st.markdown('</div>', unsafe_allow_html=True)

        # Sunburst
        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Category → Status Breakdown")
        sun_df = data.groupby(["Category","Status"]).size().reset_index(name="Count")
        fig_sun = px.sunburst(
            sun_df, path=["Category","Status"], values="Count",
            color="Status", color_discrete_map=SCLR,
        )
        fig_sun.update_traces(
            textfont=dict(size=11, color=C["text"]),
            insidetextorientation="radial",
            marker=dict(line=dict(color=C["bg"], width=2)),
        )
        fig_sun.update_layout(**_fig_base(400))
        fig_sun.update_layout(margin=dict(t=10,b=10,l=10,r=10))
        st.plotly_chart(fig_sun, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.dataframe(
            cat_df[["Category","Products","Total_Qty","At_Risk","Est. Loss (₹)"]],
            use_container_width=True, hide_index=True,
            column_config={
                "Est. Loss (₹)": st.column_config.NumberColumn(format="₹%.2f"),
                "At_Risk":       st.column_config.NumberColumn("At Risk"),
            },
        )

    # ── FINANCIAL RISK ──
    with tab_fin:
        st.markdown("<br>", unsafe_allow_html=True)
        risk_data = data[data["Status"].isin(["Expired","Critical","Expiring Soon"])].copy()

        if risk_data.empty:
            st.success("🎉  No products at risk right now!")
        else:
            total_risk = float(risk_data["Est. Loss (₹)"].sum())
            r1,r2,r3 = st.columns(3)
            with r1: kpi("At-Risk Products", len(risk_data),                     "need attention",   C["critical"])
            with r2: kpi("At-Risk Qty",      int(risk_data["Quantity"].sum()),    "total units",      C["soon"])
            with r3: kpi("Total Est. Loss",  f"₹{total_risk:,.0f}",              "potential loss",   C["expired"])

            st.markdown("<br>", unsafe_allow_html=True)

            st.markdown('<div class="glass">', unsafe_allow_html=True)
            glass_title("Estimated Loss by Category & Status")
            fig_tm = px.treemap(
                risk_data,
                path=[px.Constant("All At-Risk"),"Status","Category","Product Name"],
                values="Est. Loss (₹)",
                color="Risk Score",
                color_continuous_scale=[
                    [0.0, C["safe"]],
                    [0.4, C["soon"]],
                    [0.7, C["critical"]],
                    [1.0, C["expired"]],
                ],
                hover_data={"Quantity":True,"Days Remaining":True},
            )
            fig_tm.update_traces(
                marker=dict(line=dict(color=C["bg"],width=2)),
                textfont=dict(size=11, color=C["text"]),
            )
            fig_tm.update_layout(**_fig_base(420))
            fig_tm.update_coloraxes(showscale=False)
            fig_tm.update_layout(margin=dict(t=10,b=10,l=10,r=10))
            st.plotly_chart(fig_tm, use_container_width=True, config={"displayModeBar":False})
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            glass_title("At-Risk Product Details")
            st.dataframe(
                risk_data[["Product Name","Category","Status","Quantity",
                           "Cost Per Unit","Est. Loss (₹)","Days Remaining","Risk Score"]]
                .sort_values("Days Remaining").reset_index(drop=True),
                use_container_width=True, hide_index=True,
                column_config={
                    "Risk Score":     st.column_config.ProgressColumn(
                        "Risk Score", min_value=0, max_value=100, format="%d"),
                    "Est. Loss (₹)":  st.column_config.NumberColumn(format="₹%.2f"),
                    "Days Remaining": st.column_config.NumberColumn("Days Left", format="%d d"),
                },
            )

    # ── EXPORT ──
    with tab_dl:
        st.markdown("<br>", unsafe_allow_html=True)
        today_str = date.today().strftime("%Y%m%d")
        full_rpt = data[[
            "Product Name","Category","Batch Number","Quantity",
            "Manufacturing Date","Expiry Date","Cost Per Unit",
            "Days Remaining","Status","Risk Score","Est. Loss (₹)"
        ]].sort_values("Days Remaining").reset_index(drop=True)
        risk_rpt = full_rpt[full_rpt["Status"].isin(["Expired","Critical","Expiring Soon"])]

        st.markdown('<div class="glass">', unsafe_allow_html=True)
        glass_title("Download Options")

        d1,d2,d3 = st.columns(3)
        with d1:
            st.markdown(
                f'<div style="font-size:0.78rem;color:{C["text2"]};margin-bottom:8px">'
                f'📊 Full inventory · {len(full_rpt)} rows</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                "⬇  Full Inventory CSV",
                data=full_rpt.to_csv(index=False).encode(),
                file_name=f"inventory_{today_str}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with d2:
            st.markdown(
                f'<div style="font-size:0.78rem;color:{C["text2"]};margin-bottom:8px">'
                f'⚠️ At-risk only · {len(risk_rpt)} rows</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                "⬇  At-Risk CSV",
                data=risk_rpt.to_csv(index=False).encode(),
                file_name=f"at_risk_{today_str}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with d3:
            st.markdown(
                f'<div style="font-size:0.78rem;color:{C["text2"]};margin-bottom:8px">'
                f'📗 Styled Excel workbook · 2 sheets</div>',
                unsafe_allow_html=True,
            )
            xlsx_bytes = build_excel(full_rpt)
            st.download_button(
                "⬇  Excel Report (.xlsx)",
                data=xlsx_bytes,
                file_name=f"Expiro_Report_{today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        glass_title("Report Preview")
        st.dataframe(
            full_rpt.head(20), use_container_width=True, hide_index=True,
            column_config={
                "Risk Score":    st.column_config.ProgressColumn(
                    "Risk Score", min_value=0, max_value=100, format="%d"),
                "Est. Loss (₹)": st.column_config.NumberColumn(format="₹%.2f"),
            },
        )


# ═══════════════════════════════════════════════════════════════
# PAGE 6 ── SEARCH & FILTER
# ═══════════════════════════════════════════════════════════════
elif page == "🔍  Search & Filter":
    page_header("Search & Filter", "Find, explore, and export any subset of your inventory")

    data = gdf()

    st.markdown('<div class="glass">', unsafe_allow_html=True)
    f1,f2,f3 = st.columns([2,1.5,1.5])
    with f1:
        srch = st.text_input("Search by product name or batch number",
                             placeholder="e.g. Milk, B001…")
    with f2:
        cats    = ["All"] + sorted(data["Category"].dropna().unique().tolist())
        cat_sel = st.selectbox("Filter by category", cats)
    with f3:
        stat_sel = st.selectbox("Filter by status", ["All"]+STATUS_ORDER)
    st.markdown('</div>', unsafe_allow_html=True)

    filt = data.copy()
    if srch:
        filt = filt[
            filt["Product Name"].str.contains(srch, case=False, na=False) |
            filt["Batch Number"].str.contains(srch, case=False, na=False)
        ]
    if cat_sel  != "All": filt = filt[filt["Category"]==cat_sel]
    if stat_sel != "All": filt = filt[filt["Status"]==stat_sel]
    filt = filt.sort_values("Days Remaining").reset_index(drop=True)

    st.markdown(
        f'<div style="font-size:0.82rem;color:{C["text2"]};margin:10px 0 14px">'
        f'<b style="color:{C["accent"]}">{len(filt)}</b> product(s) found</div>',
        unsafe_allow_html=True,
    )

    if filt.empty:
        st.info("No products match your filters.")
    else:
        if len(filt) > 1:
            st.markdown('<div class="glass">', unsafe_allow_html=True)
            glass_title("Filtered Results — Days Remaining")
            fig_f = px.bar(
                filt.head(20).sort_values("Days Remaining"),
                x="Product Name", y="Days Remaining",
                color="Status", color_discrete_map=SCLR,
                text="Days Remaining",
                category_orders={"Status": STATUS_ORDER},
            )
            fig_f.update_traces(
                texttemplate="%{text}d", textposition="outside",
                textfont=dict(color=C["text2"],size=9),
                marker_line_width=0,
            )
            fig_f.update_layout(**_fig_base(240))
            fig_f.update_layout(
                bargap=0.35, xaxis_tickangle=-30,
                xaxis_title="", yaxis_title="Days Left",
                showlegend=True, legend_title_text="",
                margin=dict(t=10,b=10,l=10,r=10),
            )
            st.plotly_chart(fig_f, use_container_width=True, config={"displayModeBar":False})
            st.markdown('</div>', unsafe_allow_html=True)

        SHOW = ["Product Name","Category","Batch Number","Quantity",
                "Expiry Date","Days Remaining","Status","Risk Score","Est. Loss (₹)"]

        def row_colour(row):
            s = {
                "Safe":          f"background-color:{C['safe_bg']};color:{C['safe']}",
                "Expiring Soon": f"background-color:{C['soon_bg']};color:{C['soon']}",
                "Critical":      f"background-color:{C['crit_bg']};color:{C['critical']}",
                "Expired":       f"background-color:{C['exp_bg']};color:{C['expired']}",
            }.get(row["Status"], "")
            return [s]*len(row)

        styled = filt[SHOW].style.apply(row_colour, axis=1)
        st.dataframe(
            styled, use_container_width=True, hide_index=True,
            column_config={
                "Risk Score":     st.column_config.ProgressColumn(
                    "Risk Score", min_value=0, max_value=100, format="%d"),
                "Days Remaining": st.column_config.NumberColumn("Days Left", format="%d d"),
                "Est. Loss (₹)":  st.column_config.NumberColumn(format="₹%.2f"),
            },
        )

        st.markdown("<br>", unsafe_allow_html=True)
        dc1, dc2 = st.columns(2)
        with dc1:
            st.download_button(
                "⬇  Download Filtered CSV",
                data=filt[SHOW].to_csv(index=False).encode(),
                file_name="filtered_inventory.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with dc2:
            xlsx_f = build_excel(filt)
            st.download_button(
                "⬇  Download Filtered Excel",
                data=xlsx_f,
                file_name=f"filtered_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
